from pathlib import Path
from random import choice
from threading import Lock
from datetime import datetime, timezone

from fastapi import FastAPI, HTTPException, Request
from fastapi.responses import FileResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, Field

from career_data import DOMAINS, QUESTIONS, QUESTION_SETS, TRAIT_LABELS, get_domain_by_slug
from scoring import score_assessment


BASE_DIR = Path(__file__).resolve().parent
FEEDBACK_FILE = BASE_DIR / "feedback_responses.xlsx"
feedback_lock = Lock()
FEEDBACK_HEADERS = [
    "Timestamp",
    "Recommended Domain",
    "Recommendation Relevance",
    "Interest Level",
    "Satisfaction Rating",
    "User Comment",
]

app = FastAPI(
    title="TechPath",
    description="IT domain discovery and learning roadmap platform",
    version="1.0.0",
)
app.mount("/static", StaticFiles(directory=BASE_DIR / "static"), name="static")
templates = Jinja2Templates(directory=BASE_DIR / "templates")

ASSET_PATHS = {
    "app.css": ("css", "app.css"),
    "common.js": ("js", "common.js"),
    "domains.js": ("js", "domains.js"),
    "assessment.js": ("js", "assessment.js"),
    "results.js": ("js", "results.js"),
}


class AssessmentSubmission(BaseModel):
    answers: list[int] = Field(min_length=25, max_length=25)
    question_set_id: str = Field(min_length=1)


class FeedbackSubmission(BaseModel):
    recommended_domain: str = Field(min_length=1, max_length=120)
    recommendation_relevance: str = Field(min_length=1, max_length=40)
    interest_level: str = Field(min_length=1, max_length=40)
    satisfaction_rating: int = Field(ge=1, le=5)
    user_comment: str = Field(default="", max_length=1000)


def render(request: Request, template: str, **context):
    return templates.TemplateResponse(
        request=request,
        name=template,
        context={"domain_count": len(DOMAINS), **context},
    )


@app.get("/assets/{filename}", include_in_schema=False)
def asset(filename: str):
    relative_path = ASSET_PATHS.get(filename)
    if not relative_path:
        raise HTTPException(status_code=404, detail="Asset not found")

    flat_path = BASE_DIR / "static" / filename
    nested_path = BASE_DIR / "static" / Path(*relative_path)
    asset_path = flat_path if flat_path.is_file() else nested_path

    if not asset_path.is_file():
        raise HTTPException(status_code=404, detail="Asset not found")
    return FileResponse(asset_path)


@app.get("/", response_class=HTMLResponse)
def home(request: Request):
    return render(request, "index.html")


@app.get("/choose", response_class=HTMLResponse)
def choose_path(request: Request):
    return render(request, "choose.html")


@app.get("/domains", response_class=HTMLResponse)
def domains(request: Request):
    return render(request, "domains.html", domains=DOMAINS)


@app.get("/assessment", response_class=HTMLResponse)
def assessment(request: Request):
    return render(request, "assessment.html")


@app.get("/results", response_class=HTMLResponse)
def results(request: Request):
    return render(request, "results.html", trait_labels=TRAIT_LABELS)


@app.get("/roadmap/{slug}", response_class=HTMLResponse)
def roadmap(request: Request, slug: str):
    domain = get_domain_by_slug(slug)
    if not domain:
        raise HTTPException(status_code=404, detail="Domain not found")
    return render(request, "roadmap.html", domain=domain)


@app.get("/api/questions")
def api_questions(exclude_set_id: str | None = None):
    available_sets = list(QUESTION_SETS)
    if exclude_set_id in QUESTION_SETS and len(available_sets) > 1:
        available_sets = [set_id for set_id in available_sets if set_id != exclude_set_id]
    question_set_id = choice(available_sets)
    questions = QUESTION_SETS[question_set_id]
    return {
        "question_set_id": question_set_id,
        "questions": [
            {"id": index + 1, "text": question}
            for index, question in enumerate(questions)
        ],
        "scale": {
            "min": 1,
            "max": 5,
            "labels": ["Strongly disagree", "Disagree", "Neutral", "Agree", "Strongly agree"],
        },
    }


@app.get("/api/domains")
def api_domains():
    return {
        "domains": [
            {
                "name": domain["name"],
                "slug": domain["slug"],
                "category": domain["category"],
                "summary": domain["summary"],
                "icon": domain["icon"],
            }
            for domain in DOMAINS
        ]
    }


@app.post("/api/assessment")
def api_assessment(submission: AssessmentSubmission):
    if any(answer < 1 or answer > 5 for answer in submission.answers):
        raise HTTPException(status_code=422, detail="Every answer must be between 1 and 5")
    if submission.question_set_id not in QUESTION_SETS:
        raise HTTPException(status_code=422, detail="Unknown question set")
    return score_assessment(submission.answers, submission.question_set_id)


@app.post("/api/feedback")
def api_feedback(submission: FeedbackSubmission):
    row = [
        datetime.now(timezone.utc).isoformat(timespec="seconds"),
        submission.recommended_domain,
        submission.recommendation_relevance,
        submission.interest_level,
        submission.satisfaction_rating,
        submission.user_comment.strip(),
    ]
    print("Feedback row:")
    print("Saving to:",FEEDBACK_FILE.resolve())

    with feedback_lock:
        if FEEDBACK_FILE.exists():
            workbook = load_workbook(FEEDBACK_FILE)
            worksheet = workbook.active
            if worksheet.max_row == 0:
                worksheet.append(FEEDBACK_HEADERS)
        else:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Feedback"
            worksheet.append(FEEDBACK_HEADERS)

        worksheet.append(row)
        workbook.save(FEEDBACK_FILE)

    return {"status": "success", "message": "Feedback submitted successfully."}
    
@app.get("/download-feedback")
def download_feedback():
    if not FEEDBACK_FILE.exists():
        raise HTTPException(
            status_code=404,
            detail="No feedback file found."
        )

    return FileResponse(
        path=str(FEEDBACK_FILE),
        filename="feedback_responses.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
