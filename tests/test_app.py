from html import unescape

from fastapi.testclient import TestClient
from openpyxl import load_workbook

import app as app_module
from app import app
from career_data import DOMAINS, QUESTIONS, QUESTION_SETS


client = TestClient(app)


def test_pages_render():
    for path in ["/", "/choose", "/domains", "/assessment"]:
        response = client.get(path)
        assert response.status_code == 200


def test_frontend_assets_render_from_stable_urls():
    for filename in ["app.css", "common.js", "domains.js", "assessment.js", "results.js"]:
        response = client.get(f"/assets/{filename}")
        assert response.status_code == 200


def test_catalog_has_25_domains_and_questions():
    assert len(DOMAINS) == 25
    assert len(QUESTIONS) == 25
    assert len(QUESTION_SETS) == 5
    assert all(len(questions) == 25 for questions in QUESTION_SETS.values())
    assert len({domain["slug"] for domain in DOMAINS}) == 25


def test_questions_endpoint_returns_a_set_id():
    response = client.get("/api/questions")
    assert response.status_code == 200
    payload = response.json()
    assert payload["question_set_id"] in QUESTION_SETS
    assert len(payload["questions"]) == 25


def test_questions_endpoint_can_exclude_previous_set():
    response = client.get("/api/questions?exclude_set_id=set-1")
    assert response.status_code == 200
    payload = response.json()
    assert payload["question_set_id"] in QUESTION_SETS
    assert payload["question_set_id"] != "set-1"


def test_assessment_returns_ranked_recommendations():
    response = client.post(
        "/api/assessment",
        json={"answers": [4] * 25, "question_set_id": "set-1"},
    )
    assert response.status_code == 200
    payload = response.json()
    assert len(payload["recommendations"]) == 5
    assert len(payload["top_traits"]) == 5
    scores = [item["score"] for item in payload["recommendations"]]
    assert scores == sorted(scores, reverse=True)


def test_assessment_rejects_invalid_answers():
    response = client.post(
        "/api/assessment",
        json={"answers": [6] * 25, "question_set_id": "set-1"},
    )
    assert response.status_code == 422


def test_assessment_rejects_unknown_question_set():
    response = client.post(
        "/api/assessment",
        json={"answers": [4] * 25, "question_set_id": "missing"},
    )
    assert response.status_code == 422


def test_feedback_endpoint_appends_to_excel(monkeypatch):
    feedback_file = app_module.BASE_DIR / ".test-feedback.xlsx"
    if feedback_file.exists():
        feedback_file.unlink()
    monkeypatch.setattr(app_module, "FEEDBACK_FILE", feedback_file)

    try:
        response = client.post(
            "/api/feedback",
            json={
                "recommended_domain": "Data Analytics",
                "recommendation_relevance": "Very relevant",
                "interest_level": "Interested",
                "satisfaction_rating": 5,
                "user_comment": "Helpful shortlist.",
            },
        )

        assert response.status_code == 200
        workbook = load_workbook(feedback_file)
        worksheet = workbook.active
        assert worksheet.max_row == 2
        assert [cell.value for cell in worksheet[1]] == app_module.FEEDBACK_HEADERS
        assert worksheet["B2"].value == "Data Analytics"
        assert worksheet["E2"].value == 5
    finally:
        if feedback_file.exists():
            feedback_file.unlink()


def test_every_roadmap_renders():
    for domain in DOMAINS:
        response = client.get(f"/roadmap/{domain['slug']}")
        assert response.status_code == 200
        assert domain["name"] in unescape(response.text)
